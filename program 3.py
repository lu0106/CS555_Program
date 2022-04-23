from random import randint
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws['A1'] = 'Day of Epidemic'
ws['B1'] = 'Maximun Sick'
ws['C1'] = 'Total Immune'
ws['D1'] = 'Average Sick'
ws['E1'] = 'Sick 4 Day'
ws['F1'] = 'Sick 5 Day'
ws['G1'] = 'R'
ws['H1'] = 'Immunity'

def test(R, I):
    n = 20
    v = {}
    v[0] = int(n)       # patient number
    patient = int(n)    # patient number
    today_patient = [0]
    immune = 0      #count how many people immune
    day = 0         # day
    sick4day = 0
    sick5day = 0

    R = R * 100
    I = I * 100
    while patient >= 0 and immune <1000000:

        r = randint(1,1000000)
        v[day+1] = v[day]

        for i in range (patient):
            if patient + immune < 1000000:
                if 1 <= r <= 4:
                    for j in range (3):
                        rv = randint(1,100)
                        if 1 <= rv <= R:
                            v[day+1] = v[day+1] + 1
                            patient += 1
                    for j in range (5):
                        rv = randint(1,100)
                        if 1 <= rv <= R:
                            v[day+1] = v[day+1] + 1
                            patient += 1
                elif 5 <= r <= 12:
                    for j in range (5):
                        rv = randint(1,100)
                        if 1 <= rv <= R:
                            v[day+1] = v[day+1] + 1
                            patient += 1
                    for j in range (6):
                        rv = randint(1,100)
                        if 1 <= rv <= R:
                            v[day+1] = v[day+1] + 1
                            patient += 1
                elif 13 <= r <= 16:
                    for j in range (8):
                        rv = randint(1,100)
                        if 1 <= rv <= R:
                            v[day+1] = v[day+1] + 1
                            patient += 1
                    for j in range (7):
                        rv = randint(1,100)
                        if 1 <= rv <= R:
                            v[day+1] = v[day+1] + 1
                            patient += 1
                elif 17 <= r <= 4000:
                    for j in range (5):
                        rv = randint(1,100)
                        if 1 <= rv <= R:
                            v[day+1] = v[day+1] + 1
                            patient += 1
                    for j in range (9):
                        rv = randint(1,100)
                        if 1 <= rv <= R:
                            v[day+1] = v[day+1] + 1
                            patient += 1
                elif 4001 <= r <= 7984:
                    for j in range (8):
                        rv = randint(1,100)
                        if 1 <= rv <= R:
                            v[day+1] = v[day+1] + 1
                            patient += 1
                    for j in range (11):
                        rv = randint(1,100)
                        if 1 <= rv <= R:
                            v[day+1] = v[day+1] + 1
                            patient += 1
                else:
                    for j in range (8):
                        rv = randint(1,100)
                        if 1 <= rv <= R:
                            v[day+1] = v[day+1] + 1
                            patient += 1
                    for j in range (16):
                        rv = randint(1,100)
                        if 1 <= rv <= R:
                            v[day+1] = v[day+1] + 1
                            patient += 1

        #immune
        if day >= 1:
            for j in range (v[day-1]):
                ri = randint(1,100)
                if ri <= I:
                    immune += 1
                    v[day-1] = v[day-1] - 1
                    patient = patient -1

            if day >= 4 and v[day-4] > 0:
                sick4day = sick4day + v[day-4]

            if day >= 5 and v[day-5] > 0:
                sick5day = sick5day + v[day-5]


        day = day + 1
        today_patient.append(v[day]-v[day-1])
        if patient >=0 :
            print("Day : "+str(day)+"   Total Patient Numbe  : "+str(patient)+"   Total Immunized People : "+str(immune)+"   Today Patient Number : "+str(today_patient[day]))
    ws.append([day, max(today_patient), immune, sum(today_patient)/len(today_patient), sick4day, sick5day, R/100, I/100])


def main():

    test(0.05,0.8)
    test(0.06,0.8)
    test(0.07,0.8)
    test(0.08,0.8)
    test(0.09,0.8)
    test(0.1,0.8)
    test(0.11,0.8)
    test(0.12,0.8)
    test(0.13,0.8)
    test(0.14,0.8)
    test(0.15,0.8)

    test(0.05,0.9)
    test(0.06,0.9)
    test(0.07,0.9)
    test(0.08,0.9)
    test(0.09,0.9)
    test(0.1,0.9)
    test(0.11,0.9)
    test(0.12,0.9)
    test(0.13,0.9)
    test(0.14,0.9)
    test(0.15,0.9)

    wb.save('program 3.xlsx')

if __name__ == "__main__":
    main()